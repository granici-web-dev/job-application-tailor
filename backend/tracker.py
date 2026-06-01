from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from backend.errors import ApplicationNotFoundError, TrackerError
from backend.models import JobAnalysis
from backend.utils import sanitize_path_component

STATUS_HEADER = "Статус"
HIDDEN_HEADER = "Скрыто"
CV_FILE_HEADER = "Файл CV"
COVER_LETTER_FILE_HEADER = "Файл письма"
HEADERS = [
    "Название фирмы",
    "Сколько работников",
    "Линк на вакансию",
    "Тип занятости",
    "Название должности",
    STATUS_HEADER,
    "Дата создания",
    HIDDEN_HEADER,
    CV_FILE_HEADER,
    COVER_LETTER_FILE_HEADER,
]
STATUS_COLUMN_INDEX = HEADERS.index(STATUS_HEADER) + 1
STATUS_COLUMN = get_column_letter(STATUS_COLUMN_INDEX)
HIDDEN_COLUMN_INDEX = HEADERS.index(HIDDEN_HEADER) + 1
HIDDEN_COLUMN = get_column_letter(HIDDEN_COLUMN_INDEX)
CV_FILE_COLUMN_INDEX = HEADERS.index(CV_FILE_HEADER) + 1
COVER_LETTER_FILE_COLUMN_INDEX = HEADERS.index(COVER_LETTER_FILE_HEADER) + 1
STATUS_SENT = "ОТПРАВЛЕНО"
STATUS_NOT_SENT = "НЕ ОТПРАВЛЕНО"
STATUS_RANGE = f"{STATUS_COLUMN}2:{STATUS_COLUMN}1000"
ALLOWED_STATUSES = frozenset({STATUS_SENT, STATUS_NOT_SENT})
EMPLOYEE_COUNT_UNKNOWN = "н/д"
COLUMN_WIDTHS = {"A": 28, "B": 18, "C": 42, "D": 16, "E": 32, "F": 18, "G": 14, "H": 10, "I": 44, "J": 44}
DATA_START_ROW = 2


@dataclass(frozen=True)
class Application:
    id: int
    company_name: str
    employee_count: str
    job_url: str
    employment_type: str
    job_title: str
    status: str
    created_on: str
    hidden: bool
    cv_file: str | None
    cover_letter_file: str | None


def append_application(
    tracker_path: Path,
    analysis: JobAnalysis,
    *,
    job_url: str,
    created_on: date,
    cv_file: str | None = None,
    cover_letter_file: str | None = None,
) -> int:
    workbook = _open_or_create(tracker_path)
    sheet = workbook.active
    _ensure_schema(sheet)
    sheet.append(
        [
            analysis.company_name,
            analysis.employee_count or EMPLOYEE_COUNT_UNKNOWN,
            job_url,
            analysis.employment_type,
            analysis.job_title,
            STATUS_NOT_SENT,
            created_on.isoformat(),
            False,
            cv_file,
            cover_letter_file,
        ]
    )
    appended_row = sheet.max_row
    _save_workbook(workbook, tracker_path)
    return appended_row


def read_applications(tracker_path: Path, *, include_hidden: bool = False) -> list[Application]:
    if not tracker_path.exists():
        return []
    sheet = _load_workbook(tracker_path).active
    has_hidden_column = sheet.cell(row=1, column=HIDDEN_COLUMN_INDEX).value == HIDDEN_HEADER
    has_file_columns = sheet.cell(row=1, column=CV_FILE_COLUMN_INDEX).value == CV_FILE_HEADER
    applications = []
    for row in range(DATA_START_ROW, sheet.max_row + 1):
        values = [sheet.cell(row=row, column=column).value for column in range(1, len(HEADERS) + 1)]
        if values[0] is None:
            continue
        hidden = _parse_hidden(values[HIDDEN_COLUMN_INDEX - 1]) if has_hidden_column else False
        if hidden and not include_hidden:
            continue
        applications.append(
            Application(
                id=row,
                company_name=_cell_text(values[0]),
                employee_count=_cell_text(values[1]),
                job_url=_cell_text(values[2]),
                employment_type=_cell_text(values[3]),
                job_title=_cell_text(values[4]),
                status=_cell_text(values[5]),
                created_on=_cell_text(values[6]),
                hidden=hidden,
                cv_file=_cell_optional(values[CV_FILE_COLUMN_INDEX - 1]) if has_file_columns else None,
                cover_letter_file=_cell_optional(values[COVER_LETTER_FILE_COLUMN_INDEX - 1]) if has_file_columns else None,
            )
        )
    return applications


def set_application_status(tracker_path: Path, application_id: int, status: str) -> None:
    workbook, sheet = _open_existing_row(tracker_path, application_id)
    _ensure_schema(sheet)
    sheet.cell(row=application_id, column=STATUS_COLUMN_INDEX).value = status
    _save_workbook(workbook, tracker_path)


def set_application_hidden(tracker_path: Path, application_id: int, hidden: bool) -> None:
    workbook, sheet = _open_existing_row(tracker_path, application_id)
    _ensure_schema(sheet)
    sheet.cell(row=application_id, column=HIDDEN_COLUMN_INDEX).value = hidden
    _save_workbook(workbook, tracker_path)


def backfill_file_paths(tracker_path: Path, output_dir: Path) -> list[int]:
    applications = read_applications(tracker_path, include_hidden=True)
    company_counts = Counter(application.company_name for application in applications)
    workbook = _load_workbook(tracker_path)
    sheet = workbook.active
    _ensure_schema(sheet)
    backfilled = []
    for application in applications:
        if application.cv_file or application.cover_letter_file:
            continue
        if company_counts[application.company_name] != 1:
            continue
        folder = sanitize_path_component(application.company_name)
        folder_path = output_dir / folder
        cv_pdfs = sorted(folder_path.glob("CV_*.pdf"))
        cover_letter_pdfs = sorted(folder_path.glob("CoverLetter_*.pdf"))
        if len(cv_pdfs) != 1 or len(cover_letter_pdfs) != 1:
            continue
        sheet.cell(row=application.id, column=CV_FILE_COLUMN_INDEX).value = f"{folder}/{cv_pdfs[0].name}"
        sheet.cell(row=application.id, column=COVER_LETTER_FILE_COLUMN_INDEX).value = f"{folder}/{cover_letter_pdfs[0].name}"
        backfilled.append(application.id)
    if backfilled:
        _save_workbook(workbook, tracker_path)
    return backfilled


def _open_existing_row(tracker_path: Path, application_id: int) -> tuple[Workbook, Worksheet]:
    if not tracker_path.exists():
        raise ApplicationNotFoundError(f"No application with id {application_id} exists yet.")
    workbook = _load_workbook(tracker_path)
    sheet = workbook.active
    if application_id < DATA_START_ROW or application_id > sheet.max_row:
        raise ApplicationNotFoundError(f"No application with id {application_id} in the tracker.")
    return workbook, sheet


def _ensure_schema(sheet) -> None:
    hidden_was_missing = sheet.cell(row=1, column=HIDDEN_COLUMN_INDEX).value != HIDDEN_HEADER
    for index, header in enumerate(HEADERS, start=1):
        header_cell = sheet.cell(row=1, column=index)
        if header_cell.value == header:
            continue
        header_cell.value = header
        header_cell.font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS[get_column_letter(index)]
    if hidden_was_missing:
        for row in range(DATA_START_ROW, sheet.max_row + 1):
            sheet.cell(row=row, column=HIDDEN_COLUMN_INDEX).value = False


def _parse_hidden(value: object) -> bool:
    if value is True:
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "да", "yes"}
    return False


def _cell_text(value: object) -> str:
    return "" if value is None else str(value)


def _cell_optional(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _save_workbook(workbook: Workbook, tracker_path: Path) -> None:
    try:
        workbook.save(tracker_path)
    except PermissionError:
        raise TrackerError(
            f"Could not write the tracker at {tracker_path}. It is probably open in Excel. "
            "Close it and run again."
        )


def _open_or_create(tracker_path: Path) -> Workbook:
    if not tracker_path.exists():
        return _create_workbook()
    return _load_workbook(tracker_path)


def _load_workbook(tracker_path: Path) -> Workbook:
    try:
        return load_workbook(tracker_path)
    except PermissionError:
        raise TrackerError(
            f"Could not open the tracker at {tracker_path}. It is probably open in Excel. "
            "Close it and run again."
        )


def _create_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applications"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    status_dropdown = DataValidation(
        type="list",
        formula1=f'"{STATUS_SENT},{STATUS_NOT_SENT}"',
        allow_blank=False,
    )
    sheet.add_data_validation(status_dropdown)
    status_dropdown.add(STATUS_RANGE)
    return workbook
