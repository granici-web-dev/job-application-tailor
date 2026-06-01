import json
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend import api
from backend.api import Paths, app, get_anthropic_client, get_paths, get_settings, serve_file
from backend.config import Settings
from backend.errors import TrackerError
from backend.models import JobAnalysis
from backend.pipeline import TRACKER_FILENAME
from backend.tracker import STATUS_NOT_SENT, STATUS_SENT, append_application

PROFILE_DIR = Path(__file__).parent / "fixtures" / "profile"

ANALYSIS_JSON = json.dumps(
    {
        "company_name": "Acme GmbH",
        "job_title": "Senior Backend Engineer",
        "employment_type": "remote",
        "employee_count": None,
        "hard_skills": ["Python", "FastAPI"],
        "keywords": ["Python"],
    }
)
CV_MARKDOWN = "# Ivan Petrov\n\n## Summary\nSenior Backend Engineer with Python and FastAPI."
LETTER_MARKDOWN = "Dear Acme GmbH team,\n\nKind regards,\nIvan"


def _message(text):
    return SimpleNamespace(content=[SimpleNamespace(type="text", text=text)])


def _fake_client():
    client = Mock()
    client.messages.create.side_effect = [
        _message(ANALYSIS_JSON),
        _message(CV_MARKDOWN),
        _message(CV_MARKDOWN),
        _message(LETTER_MARKDOWN),
        _message(LETTER_MARKDOWN),
    ]
    return client


def _seed_application(
    output_dir: Path,
    company: str,
    cv_file: str | None = None,
    cover_letter_file: str | None = None,
) -> int:
    analysis = JobAnalysis.from_response_dict(
        {"company_name": company, "job_title": "Engineer", "employment_type": "remote"}
    )
    return append_application(
        output_dir / TRACKER_FILENAME,
        analysis,
        job_url=f"https://example.com/{company}",
        created_on=date(2026, 6, 1),
        cv_file=cv_file,
        cover_letter_file=cover_letter_file,
    )


@pytest.fixture
def client(tmp_path):
    app.dependency_overrides[get_paths] = lambda: Paths(output_dir=tmp_path, profile_dir=PROFILE_DIR)
    with TestClient(app) as test_client:
        test_client.output_dir = tmp_path
        yield test_client
    app.dependency_overrides.clear()


def test_generate_returns_ids_and_serves_files(client):
    app.dependency_overrides[get_settings] = lambda: Settings(anthropic_api_key="test", model="claude-sonnet-4-6")
    app.dependency_overrides[get_anthropic_client] = _fake_client

    response = client.post("/generate", json={"url": "https://example.com/jobs/1", "text": "Backend role at Acme GmbH."})

    assert response.status_code == 200
    body = response.json()
    assert body["id"] == 2
    assert body["company"] == "Acme_GmbH"
    assert body["analysis"]["company_name"] == "Acme GmbH"
    assert body["files"]["cv"] == f"/files/Acme_GmbH/{body['cv_pdf']}"

    sheet = load_workbook(client.output_dir / TRACKER_FILENAME).active
    assert sheet["A2"].value == "Acme GmbH"

    pdf = client.get(body["files"]["cv"])
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")


def test_list_applications_returns_rows_with_ids(client):
    _seed_application(client.output_dir, "Acme")
    _seed_application(client.output_dir, "Globex")

    rows = client.get("/applications").json()

    assert [row["id"] for row in rows] == [2, 3]
    assert [row["company_name"] for row in rows] == ["Acme", "Globex"]
    assert rows[0]["status"] == STATUS_NOT_SENT


def test_list_applications_empty_when_no_tracker(client):
    assert client.get("/applications").json() == []


def test_patch_status_syncs_to_excel(client):
    row = _seed_application(client.output_dir, "Acme")

    response = client.patch(f"/applications/{row}/status", json={"status": STATUS_SENT})

    assert response.status_code == 200
    assert response.json()["status"] == STATUS_SENT
    sheet = load_workbook(client.output_dir / TRACKER_FILENAME).active
    assert sheet[f"F{row}"].value == STATUS_SENT


def test_patch_status_rejects_unknown_value(client):
    row = _seed_application(client.output_dir, "Acme")

    response = client.patch(f"/applications/{row}/status", json={"status": "MAYBE"})

    assert response.status_code == 422


def test_patch_status_unknown_id_returns_404(client):
    _seed_application(client.output_dir, "Acme")

    response = client.patch("/applications/99/status", json={"status": STATUS_SENT})

    assert response.status_code == 404


def test_patch_status_locked_tracker_returns_409(client, monkeypatch):
    row = _seed_application(client.output_dir, "Acme")

    def _locked(*args, **kwargs):
        raise TrackerError("It is probably open in Excel.")

    monkeypatch.setattr(api, "set_application_status", _locked)
    response = client.patch(f"/applications/{row}/status", json={"status": STATUS_SENT})

    assert response.status_code == 409


def test_hide_removes_application_from_active_list(client):
    row = _seed_application(client.output_dir, "Acme")

    assert client.patch(f"/applications/{row}/visibility", json={"hidden": True}).status_code == 200

    active = client.get("/applications").json()
    assert all(application["id"] != row for application in active)


def test_include_hidden_returns_hidden_applications(client):
    row = _seed_application(client.output_dir, "Acme")
    client.patch(f"/applications/{row}/visibility", json={"hidden": True})

    rows = client.get("/applications", params={"include_hidden": "true"}).json()
    target = next(application for application in rows if application["id"] == row)
    assert target["hidden"] is True


def test_application_response_includes_hidden_flag(client):
    _seed_application(client.output_dir, "Acme")

    rows = client.get("/applications").json()
    assert rows[0]["hidden"] is False


def test_patch_visibility_idempotent_hide(client):
    row = _seed_application(client.output_dir, "Acme")

    first = client.patch(f"/applications/{row}/visibility", json={"hidden": True})
    second = client.patch(f"/applications/{row}/visibility", json={"hidden": True})

    assert first.status_code == 200 and second.status_code == 200
    assert second.json()["hidden"] is True
    everything = client.get("/applications", params={"include_hidden": "true"}).json()
    assert sum(1 for application in everything if application["id"] == row) == 1


def test_show_again_returns_to_active_list(client):
    row = _seed_application(client.output_dir, "Acme")
    client.patch(f"/applications/{row}/visibility", json={"hidden": True})

    client.patch(f"/applications/{row}/visibility", json={"hidden": False})

    active = client.get("/applications").json()
    assert any(application["id"] == row for application in active)


def test_patch_visibility_unknown_id_returns_404(client):
    _seed_application(client.output_dir, "Acme")

    assert client.patch("/applications/99/visibility", json={"hidden": True}).status_code == 404


def test_applications_include_file_links(client):
    _seed_application(
        client.output_dir,
        "Acme",
        cv_file="Acme/CV_Ivan_Acme.pdf",
        cover_letter_file="Acme/CoverLetter_Ivan_Acme.pdf",
    )

    row = client.get("/applications").json()[0]
    assert row["files"] == {
        "cv": "/files/Acme/CV_Ivan_Acme.pdf",
        "cover_letter": "/files/Acme/CoverLetter_Ivan_Acme.pdf",
    }


def test_collision_rows_have_distinct_file_links(client):
    _seed_application(client.output_dir, "Jobriver", cv_file="Jobriver/CV_Ivan_Jobriver.pdf", cover_letter_file="Jobriver/CoverLetter_Ivan_Jobriver.pdf")
    _seed_application(client.output_dir, "Jobriver", cv_file="Jobriver_2/CV_Ivan_Jobriver.pdf", cover_letter_file="Jobriver_2/CoverLetter_Ivan_Jobriver.pdf")

    rows = client.get("/applications").json()
    cv_links = {row["files"]["cv"] for row in rows}
    assert cv_links == {"/files/Jobriver/CV_Ivan_Jobriver.pdf", "/files/Jobriver_2/CV_Ivan_Jobriver.pdf"}


def test_legacy_row_returns_null_file_links(client):
    _seed_application(client.output_dir, "Acme")

    row = client.get("/applications").json()[0]
    assert row["files"] is None


def test_files_missing_returns_404(client):
    assert client.get("/files/Nope/missing.pdf").status_code == 404


def test_serve_file_rejects_path_traversal(tmp_path):
    secret = tmp_path / "secret.txt"
    secret.write_text("private")
    output_dir = tmp_path / "output"
    output_dir.mkdir()

    with pytest.raises(HTTPException) as raised:
        serve_file(company="..", filename="secret.txt", paths=Paths(output_dir=output_dir, profile_dir=PROFILE_DIR))

    assert raised.value.status_code == 404
