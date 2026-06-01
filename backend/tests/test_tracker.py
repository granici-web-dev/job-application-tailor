from datetime import date
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from backend.errors import ApplicationNotFoundError, TrackerError
from backend.models import JobAnalysis
from backend.tracker import (
    COVER_LETTER_FILE_HEADER,
    CV_FILE_HEADER,
    EMPLOYEE_COUNT_UNKNOWN,
    HEADERS,
    HIDDEN_HEADER,
    LOCATION_HEADER,
    STATUS_NOT_SENT,
    STATUS_SENT,
    append_application,
    backfill_file_paths,
    read_applications,
    set_application_hidden,
)

LEGACY_HEADERS = [
    "Название фирмы",
    "Сколько работников",
    "Линк на вакансию",
    "Тип занятости",
    "Название должности",
    "Статус",
    "Дата создания",
]


def _legacy_tracker(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applications"
    sheet.append(LEGACY_HEADERS)
    sheet.append(["Acme GmbH", "50-200", "https://x/1", "remote", "Backend Engineer", STATUS_NOT_SENT, "2026-05-01"])
    dropdown = DataValidation(type="list", formula1=f'"{STATUS_SENT},{STATUS_NOT_SENT}"', allow_blank=False)
    sheet.add_data_validation(dropdown)
    dropdown.add("F2:F1000")
    workbook.save(path)


def _analysis(**overrides):
    data = {
        "company_name": "Acme GmbH",
        "job_title": "Backend Engineer",
        "employment_type": "remote",
        "employee_count": "50-200",
    }
    data.update(overrides)
    return JobAnalysis.from_response_dict(data)


def test_append_creates_file_with_header_when_missing(tmp_path):
    path = tmp_path / "tracker.xlsx"

    append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))

    assert path.exists()
    sheet = load_workbook(path).active
    assert [cell.value for cell in sheet[1]] == HEADERS
    assert sheet[1][0].font.bold is True
    assert sheet.freeze_panes == "A2"


def test_append_adds_row_with_values_and_na_employee_count(tmp_path):
    path = tmp_path / "tracker.xlsx"

    row = append_application(path, _analysis(employee_count=None), job_url="https://x/1", created_on=date(2026, 6, 1))

    assert row == 2
    sheet = load_workbook(path).active
    assert [cell.value for cell in sheet[2]] == [
        "Acme GmbH",
        EMPLOYEE_COUNT_UNKNOWN,
        "https://x/1",
        "remote",
        "Backend Engineer",
        STATUS_NOT_SENT,
        "2026-06-01",
        False,
        None,
        None,
        None,
    ]


def test_append_appends_without_overwriting(tmp_path):
    path = tmp_path / "tracker.xlsx"

    append_application(path, _analysis(company_name="First"), job_url="https://x/1", created_on=date(2026, 6, 1))
    second = append_application(path, _analysis(company_name="Second"), job_url="https://x/2", created_on=date(2026, 6, 2))

    assert second == 3
    sheet = load_workbook(path).active
    assert sheet["A2"].value == "First"
    assert sheet["A3"].value == "Second"
    assert sheet.max_row == 3


def test_status_column_dropdown_is_bound_to_the_status_column(tmp_path):
    path = tmp_path / "tracker.xlsx"
    append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))
    sheet = load_workbook(path).active

    columns = {cell.value: cell.column_letter for cell in sheet[1]}
    assert columns["Статус"] == "F"
    assert columns[HIDDEN_HEADER] == "H"
    assert columns[CV_FILE_HEADER] == "I"
    assert columns[COVER_LETTER_FILE_HEADER] == "J"
    assert columns[LOCATION_HEADER] == "K"

    status_dropdowns = [
        dv for dv in sheet.data_validations.dataValidation if dv.formula1 and "ОТПРАВЛЕНО" in dv.formula1
    ]
    assert len(status_dropdowns) == 1
    assert str(status_dropdowns[0].sqref) == "F2:F1000"
    assert "НЕ ОТПРАВЛЕНО" in status_dropdowns[0].formula1


def test_new_rows_default_to_not_hidden(tmp_path):
    path = tmp_path / "tracker.xlsx"
    append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))

    applications = read_applications(path)
    assert len(applications) == 1
    assert applications[0].hidden is False


def test_read_excludes_hidden_by_default_and_includes_on_request(tmp_path):
    path = tmp_path / "tracker.xlsx"
    visible = append_application(path, _analysis(company_name="Visible"), job_url="https://x/1", created_on=date(2026, 6, 1))
    hidden = append_application(path, _analysis(company_name="Hidden"), job_url="https://x/2", created_on=date(2026, 6, 1))
    set_application_hidden(path, hidden, True)

    active = read_applications(path)
    assert [a.id for a in active] == [visible]

    everything = read_applications(path, include_hidden=True)
    assert [a.id for a in everything] == [visible, hidden]
    assert next(a for a in everything if a.id == hidden).hidden is True


def test_set_application_hidden_marks_and_unmarks_row(tmp_path):
    path = tmp_path / "tracker.xlsx"
    row = append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))

    set_application_hidden(path, row, True)
    assert read_applications(path, include_hidden=True)[0].hidden is True

    set_application_hidden(path, row, False)
    assert read_applications(path, include_hidden=True)[0].hidden is False


def test_set_application_hidden_is_idempotent(tmp_path):
    path = tmp_path / "tracker.xlsx"
    row = append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))

    set_application_hidden(path, row, True)
    set_application_hidden(path, row, True)

    everything = read_applications(path, include_hidden=True)
    assert len(everything) == 1
    assert everything[0].hidden is True


def test_set_application_hidden_unknown_id_raises(tmp_path):
    path = tmp_path / "tracker.xlsx"
    append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))

    with pytest.raises(ApplicationNotFoundError):
        set_application_hidden(path, 99, True)


def test_legacy_tracker_without_hidden_column_backfilled_on_write(tmp_path):
    path = tmp_path / "tracker.xlsx"
    _legacy_tracker(path)

    append_application(path, _analysis(company_name="New Co"), job_url="https://x/2", created_on=date(2026, 6, 1))

    sheet = load_workbook(path).active
    assert sheet.cell(row=1, column=8).value == HIDDEN_HEADER
    assert sheet.cell(row=2, column=8).value is False
    assert sheet.cell(row=3, column=8).value is False
    assert sheet.cell(row=1, column=9).value == CV_FILE_HEADER
    assert sheet.cell(row=1, column=10).value == COVER_LETTER_FILE_HEADER
    assert sheet.cell(row=1, column=11).value == LOCATION_HEADER

    status_dropdowns = [
        dv for dv in sheet.data_validations.dataValidation if dv.formula1 and "ОТПРАВЛЕНО" in dv.formula1
    ]
    assert len(status_dropdowns) == 1
    assert str(status_dropdowns[0].sqref) == "F2:F1000"
    assert next(cell.column_letter for cell in sheet[1] if cell.value == "Статус") == "F"


def test_read_legacy_tracker_treats_rows_as_active(tmp_path):
    path = tmp_path / "tracker.xlsx"
    _legacy_tracker(path)

    applications = read_applications(path)
    assert len(applications) == 1
    assert applications[0].hidden is False
    assert applications[0].company_name == "Acme GmbH"


def test_append_stores_and_reads_file_paths(tmp_path):
    path = tmp_path / "tracker.xlsx"
    append_application(
        path,
        _analysis(),
        job_url="https://x/1",
        created_on=date(2026, 6, 1),
        cv_file="Acme_GmbH/CV_Ivan_Acme_GmbH.pdf",
        cover_letter_file="Acme_GmbH/CoverLetter_Ivan_Acme_GmbH.pdf",
    )

    application = read_applications(path)[0]
    assert application.cv_file == "Acme_GmbH/CV_Ivan_Acme_GmbH.pdf"
    assert application.cover_letter_file == "Acme_GmbH/CoverLetter_Ivan_Acme_GmbH.pdf"


def test_read_legacy_row_has_no_file_paths(tmp_path):
    path = tmp_path / "tracker.xlsx"
    _legacy_tracker(path)

    application = read_applications(path)[0]
    assert application.cv_file is None
    assert application.cover_letter_file is None


def test_append_stores_and_reads_location(tmp_path):
    path = tmp_path / "tracker.xlsx"
    append_application(
        path,
        _analysis(),
        job_url="https://x/1",
        created_on=date(2026, 6, 1),
        location="Wien, Österreich",
    )

    assert read_applications(path)[0].location == "Wien, Österreich"


def test_read_legacy_row_has_no_location(tmp_path):
    path = tmp_path / "tracker.xlsx"
    _legacy_tracker(path)

    assert read_applications(path)[0].location is None


def _company_folder_with_pdfs(output_dir, folder, full_name):
    company_dir = output_dir / folder
    company_dir.mkdir(parents=True)
    (company_dir / f"CV_{full_name}_{folder}.pdf").write_bytes(b"%PDF cv")
    (company_dir / f"CoverLetter_{full_name}_{folder}.pdf").write_bytes(b"%PDF letter")


def test_backfill_sets_link_for_unique_company_and_skips_collisions(tmp_path):
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    path = output_dir / "tracker.xlsx"

    unique = append_application(path, _analysis(company_name="Weissenberg Group GmbH"), job_url="https://x/1", created_on=date(2026, 6, 1))
    first_collision = append_application(path, _analysis(company_name="Jobriver HR Service"), job_url="https://x/2", created_on=date(2026, 6, 1))
    second_collision = append_application(path, _analysis(company_name="Jobriver HR Service"), job_url="https://x/3", created_on=date(2026, 6, 1))

    _company_folder_with_pdfs(output_dir, "Weissenberg_Group_GmbH", "Serghei_Granici")
    _company_folder_with_pdfs(output_dir, "Jobriver_HR_Service", "Serghei_Granici")
    _company_folder_with_pdfs(output_dir, "Jobriver_HR_Service_2", "Serghei_Granici")

    backfilled = backfill_file_paths(path, output_dir)

    assert backfilled == [unique]
    by_id = {a.id: a for a in read_applications(path, include_hidden=True)}
    assert by_id[unique].cv_file == "Weissenberg_Group_GmbH/CV_Serghei_Granici_Weissenberg_Group_GmbH.pdf"
    assert by_id[unique].cover_letter_file == "Weissenberg_Group_GmbH/CoverLetter_Serghei_Granici_Weissenberg_Group_GmbH.pdf"
    assert by_id[first_collision].cv_file is None
    assert by_id[second_collision].cv_file is None


def test_append_default_status_is_not_sent(tmp_path):
    path = tmp_path / "tracker.xlsx"
    append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))
    sheet = load_workbook(path).active

    assert sheet["F2"].value == STATUS_NOT_SENT


def test_append_raises_tracker_error_when_file_is_locked(tmp_path):
    path = tmp_path / "tracker.xlsx"

    with patch("backend.tracker.Workbook.save", side_effect=PermissionError):
        with pytest.raises(TrackerError, match="open in Excel"):
            append_application(path, _analysis(), job_url="https://x/1", created_on=date(2026, 6, 1))
